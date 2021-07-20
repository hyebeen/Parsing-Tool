# -*- coding: utf-8 -*-
# PYTHON 3.6, 3.7, 3.9

# pyinstaller --icon=pngegg.ico --onefile ParsingTool.py

# 개선할 점
# 1. 양호취약 판별 로직 수정
# 2. utf-8 변환 에러 수정
# 3. 페이지 나누기 미리보기

import io, os, time, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.views import Pane

def ConvertFile(pathDir, fileList) :
    print("\n*** 파싱을 위해 대상 파일을 utf-8로 변환 실행")
    for fileName in fileList:
        fileDir = pathDir + '\\' + fileName
        try:
            file = io.open(fileDir, mode='r').read()
        except UnicodeError:
            continue
        io.open(fileDir, mode='w', encoding="utf-8").write(file)
        print("\n*** utf-8 포맷으로 변환 완료")

def InputInformation():
    pathDir = input("\n* 파싱 대상 파일이 있는 경로 입력\n  Ex) C:\\Users\hyebeen\\Desktop\\directory \n  => ")
    try :
        fileList = os.listdir(pathDir)
    except FileNotFoundError :
        print("*** 잘못된 경로")
        time.sleep(2)
        sys.exit()

    startLine = input("\n* \"시작 패턴\"을 입력\n (default)##### START ##### \n  =>  ")
    if startLine == "" : startLine = "##### START #####"

    newStartLine = ""
    startString = input("\n* 파싱 결과에 시작 문자열을 추가\n  Ex) 예:y,  (default)아니오:n \n  => ")
    if startString == "" : startString = "n"
    elif startString == "y" :
        newStartLine = input("\n* 파싱 결과에 추가 할 시작 문자열을 입력\n (default)[시스템 현황]\n =>  ")
        newStartLine += "\n"
        if newStartLine == "" : newStartLine = "[시스템 현황]\n"

    endLine = input("\n* \"마지막 패턴\"을 입력\n  Ex) (default)##### END ##### \n  =>  ")
    if endLine == "" : endLine = "##### END #####"

    resultLine = ""
    resultFlag = input("\n* 양호, 취약 여부 파싱\n* 스크립트 결과 파일에 양호, 취약여부가 나와야 함.\n  예:y,  (default)아니오:n \n  =>  ")
    if resultFlag == "y" :
        resultFlag = 2
        resultLine = input("\n* \"결과 패턴\" 입력\n  Ex) \"★ U_01. 결과 : 양호\" 이면 \"결과 : \" 입력\n* \"는 제외하고 입력)\n  =>  ")
    elif resultFlag == "n" or resultFlag == "" :
        resultFlag = 1

    return pathDir, startLine, newStartLine, endLine, fileList, resultFlag, resultLine


def ResultParse(ws, txt, resultLine, rowIndex, colIndex) :
    if resultLine in txt:
        if "양호" in txt or "O" in txt or "o" in txt:
            ws.cell(rowIndex, colIndex - 1, "O")
            # ws.cell(rowIndex, colIndex - 1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA',
            #                                                    fill_type='solid')
            ws.cell(rowIndex, colIndex - 1).alignment = Alignment(horizontal='center', vertical='center')
            # ws.cell(rowIndex, colIndex - 1).font = Font(color='76933C')
        elif "취약" in txt or "X" in txt or "x" in txt:
            ws.cell(rowIndex, colIndex - 1, "X")
            ws.cell(rowIndex, colIndex - 1).fill = PatternFill(start_color='FFD9D9', end_color='FFD9D9',
                                                               fill_type='solid')
            ws.cell(rowIndex, colIndex - 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(rowIndex, colIndex - 1).font = Font(color='FF0000')
        else:
            ws.cell(rowIndex, colIndex - 1, txt)


def SetStyle(resultFlag, ws, rowIndex):
    ScolIndex = 11
    SrowIndex = rowIndex
    if resultFlag == 2 : ScolIndex -= 1

    # 행 높이 설정
    for row in range(3, rowIndex + 1):
        ws.row_dimensions[row].height = 18

    # 열 너비 설정
    ws.column_dimensions['J'].width = 47

    # 기본 틀 스타일 설정
    # 제목 글자
    ws.cell(1, 1).font = Font(bold=True, size=11, color='000000')

    # 표 제목 글자
    for r in range(1, ScolIndex+1):
        ws.cell(3, r).font = Font(bold=True, size=9, color='002060')
        ws.cell(3, r).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(3, r).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 표 내용 글자
    for i in range(4, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            ws.cell(i, s).font = Font(size=9, color='000000')


    # 테두리 설정
    # 표 내용 테두리
    for i in range(4, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            ws.cell(i, s).border = Border(left =Side(style='thin', color='D9D9D9'),
                                            right = Side(style='thin', color='D9D9D9'),
                                            top=Side(style='thin', color='D9D9D9'),
                                            bottom=Side(style='thin', color='D9D9D9'))
    # 표 제목 테두리
    for s in range(1, ScolIndex+1):
        ws.cell(3, s).border = Border(left =Side(style='thin', color='D9D9D9'),
                                        right = Side(style='thin', color='D9D9D9'),
                                        top=Side(style='medium', color='002060'),
                                        bottom=Side(style='thin', color='002060'))

def RemoveFile() :
    if os.path.isfile("상세결과.xlsx"):
        removeFlag = input("\n*** 기존 파일(\"상세결과.xlsx\")이 존재합니다.\n*** 계속 진행하려면 삭제해야 합니다. 삭제하시겠습니까?\n  예:y, 아니오:n\n =>  ")
        if removeFlag == "y" :
            try:
                os.remove("상세결과.xlsx")
            except PermissionError:
                print("\n*** \"상세결과.xlsx\"파일을 닫은 후 재실행해주세요.")
                time.sleep(2)
                sys.exit()
            print("\n*** 기존 파일(\"상세결과.xlsx\")을 삭제하였습니다.")
        if removeFlag == "n":
            print("기존 파일(\"상세결과.xlsx\") 삭제 후 실행해야 합니다.")
            time.sleep(2)
            sys.exit()


def StartParse(pathDir, startLine, newStartLine, endLine, fileList, resultFlag, resultLine) :
    # 엑셀 시트 및 기본 틀 생성
    wb = Workbook()
    ws = wb.active

    ws.sheet_view.showGridLines = False

    #ws.sheet_properties.pageSetUpPr = PageSetupProperties(autoPageBreaks=True, fitToPage=True)
    #ws.sheet_properties.pageSetUpPr.autoPageBreaks=True

    ws.title = "상세결과"

    ws.cell(1, 1, "■ 상세결과")
    ws.cell(3, 1, "점검영역")
    ws.cell(3, 2, "CODE")
    ws.cell(3, 3, "점검항목")
    ws.cell(3, 4, "위험도")
    ws.cell(3, 5, "자산명")
    ws.cell(3, 6, "IP")
    ws.cell(3, 7, "분류")
    ws.cell(3, 8, "진단결과")
    ws.cell(3, 9, "조치결과")
    ws.cell(3, 10, "현재설정")
    ws.cell(3, 11, "비고")

    readFlag = "False"
    write_txt = newStartLine
    rowIndex = 3
    colIndex = 10
    resultFile = "상세결과.xlsx"

    # 파일 하나씩 열어서 파싱
    for fileName in fileList:
        fileDir = pathDir + '\\' + fileName
        txt_f = open(fileDir, encoding="utf-8", mode="r")

        while True:
            txt = txt_f.readline()
            if startLine in txt :
                readFlag = "True"
                rowIndex += 1
                continue
            if endLine in txt :
                readFlag = "False"
                ws.cell(rowIndex, 5, fileName)
                ws.cell(rowIndex, 6, fileName)
                ws.cell(rowIndex, 7, fileName)
                ws.cell(rowIndex, colIndex, write_txt)
                ws.cell(rowIndex, colIndex).alignment = Alignment(wrap_text=True)
                write_txt = newStartLine
            if readFlag == "True":
                write_txt += txt
            if not txt: break

        txt_f.close()

    # 스타일 설정
    SetStyle(resultFlag, ws, rowIndex)

    # 파일 저장
    wb.save(resultFile)


def main():
    # 기존 파일 있으면 삭제
    RemoveFile()

    # 정보 입력
    pathDir, startLine, newStartLine, endLine, fileList, resultFlag, resultLine = InputInformation()

    # utf-8로 변환
    ConvertFile(pathDir, fileList)

    # 파싱 실행
    StartParse(pathDir, startLine, newStartLine, endLine, fileList, resultFlag, resultLine)

    # 종료
    print("\n*** 완료 ***\n*** \"파싱 프로그램\"이 있는 경로에 \"상세결과.xlsx\" 결과 파일 생성 ***")
    time.sleep(5)


if __name__ == '__main__':
    main()
