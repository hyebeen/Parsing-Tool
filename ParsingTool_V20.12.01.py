# -*- coding: utf-8 -*-
# PYTHON 3.6, 3.7, 3.9

# pyinstaller --icon=pngegg.ico --onefile ParsingTool.py

# 개선할 점
# 1. 문자열 입력 -> try except로 검사하여 재입력받기

import io, os, time, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def ConvertFile(pathDir, fileList) :
    print("\n*** 파싱을 위해 대상 파일을 utf-8로 변환합니다.")
    for fileName in fileList:
        fileDir = pathDir + '\\' + fileName
        try:
            file = io.open(fileDir, mode='r').read()
        except UnicodeError:
            continue
        io.open(fileDir, mode='w', encoding="utf-8").write(file)
        print("\n*** utf-8 포맷으로 변환이 완료되었습니다.")

def InputInformation():
    pathDir = input("\n* 파싱 대상 파일이 있는 디렉터리를 입력하세요.\n  Ex) C:\\Users\hyebeen\\Desktop\\directory \n  => ")
    try :
        fileList = os.listdir(pathDir)
    except FileNotFoundError :
        print("*** 입력하신 경로가 존재하지 않습니다.")
        time.sleep(2)
        sys.exit()
    startLine = input("\n* \"시작 패턴\"을 입력하세요.\n  Ex) ##### START ##### \n  =>  ")
    startFlag = input("\n* \"시작 패턴\"을 포함하여 파싱하시겠습니까?\n  예 : y, 아니오 : n \n  =>  ")
    newStartLine = ""
    if startFlag == "n" :
        if input("\n* 파싱 결과에 시작 문자열을 추가하시겠습니까?  Ex) [시스템 현황]\n  예 : y, 아니오 : n \n  => ") == "y" :
            newStartLine = input("\n* 파싱 결과에 추가 할 시작 문자열을 입력하세요.  =>  ")
            newStartLine += "\n"
    endLine = input("\n* \"마지막 패턴\"을 입력하세요. \n  Ex) ##### END ##### \n  =>  ")
    endFlag = input("\n* \"마지막 패턴\"을 포함하여 파싱하시겠습니까?\n  예 : y, 아니오 : n \n  =>  ")
    resultLine = ""
    resultFlag = input("\n* 양호, 취약 여부를 파싱하시겠습니까?\n* 스크립트 결과 파일에 양호, 취약여부가 나와야 합니다.\n  예 : y, 아니오 : n \n  =>  ")
    if resultFlag == "y" :
        resultFlag = 2
        resultLine = input("\n* \"결과 패턴\"을 입력하세요.\n  Ex) \"★ U_01. 결과 : 양호\" 이면 \"결과 : \" 입력\n* \"는 제외하고 입력)\n  =>  ")
    elif resultFlag =="n" :
        resultFlag = 1

    return pathDir, startLine, startFlag, newStartLine, endLine, endFlag, fileList, resultFlag, resultLine


def ResultParse(ws, txt, resultLine, rowIndex, colIndex) :
    if resultLine in txt:
        if "양호" in txt or "O" in txt or "o" in txt:
            ws.cell(rowIndex, colIndex - 1, "O")
            # ws.cell(rowIndex, colIndex - 1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA',
            #                                                    fill_type='solid')
            ws.cell(rowIndex, colIndex - 1).alignment = Alignment(horizontal='center', vertical='center')
            # ws.cell(rowIndex, colIndex - 1).font = Font(color='76933C')
        elif "취약" in txt or "X" in txt or "x" in txt:
            ws.cell(rowIndex, colIndex - 1, "X")
            ws.cell(rowIndex, colIndex - 1).fill = PatternFill(start_color='FFD9D9', end_color='FFD9D9',
                                                               fill_type='solid')
            ws.cell(rowIndex, colIndex - 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(rowIndex, colIndex - 1).font = Font(color='FF0000')
        else:
            ws.cell(rowIndex, colIndex - 1, txt)


def SetStyle(resultFlag, ws, colIndex, rowIndex):
    if resultFlag == 2 : colIndex -= 1

    # 행 높이 설정
    for row in range(2, rowIndex + 1):
        ws.row_dimensions[row].height = 30

    # 기본 틀 스타일 설정
    for r in range(2, colIndex):
        ws.cell(2, r).font = Font(bold=True)
        ws.cell(2, r).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(2, r).fill = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')

    # 테두리 설정
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for i in range(2, rowIndex + 1):
        for s in range(2, colIndex):
            ws.cell(i, s).border = thin_border


def RemoveFile() :
    if os.path.isfile("ParsingResult.xlsx"):
        removeFlag = input("\n*** 기존 파일(\"ParsingResult.xlsx\")이 존재합니다.\n*** 계속 진행하려면 삭제해야 합니다. 삭제하시겠습니까?\n  예 :y, 아니오 : n\n =>  ")
        if removeFlag == "y" :
            try:
                os.remove("ParsingResult.xlsx")
            except PermissionError:
                print("\n*** \"ParsingResult.xlsx\"파일을 닫은 후 재실행해주세요.")
                time.sleep(2)
                sys.exit()
            print("\n*** 기존 파일(\"ParsingResult.xlsx\")을 삭제하였습니다.")
        if removeFlag == "n":
            print("기존 파일(\"ParsingResult.xlsx\") 삭제 후 실행해야 합니다.")
            time.sleep(2)
            sys.exit()


def StartParse(pathDir, startLine, startFlag, newStartLine, endLine, endFlag, fileList, resultFlag, resultLine) :
    # 엑셀 시트 및 기본 틀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    ws.cell(2, 2, "항목")
    ws.cell(2, 3, "설명")

    readFlag = "False"
    write_txt = newStartLine
    colIndex = 5 if resultFlag == 2 else 4
    resultFile = "ParsingResult.xlsx"

    # 파일 하나씩 열어서 파싱
    for fileName in fileList:
        fileDir = pathDir + '\\' + fileName
        txt_f = open(fileDir, encoding="utf-8", mode="r")

        rowIndex = 2
        if resultFlag == 2 :
            ws.merge_cells(start_row=rowIndex, start_column=colIndex - 1, end_row=rowIndex, end_column=colIndex)
            ws.cell(rowIndex, colIndex - 1, fileName)
        elif resultFlag == 1 :
            ws.cell(rowIndex, colIndex, fileName)

        while True:
            txt = txt_f.readline()
            if startLine in txt :
                readFlag = "True"
                rowIndex += 1
                if startFlag == "n" : continue
            if resultFlag == 2 : ResultParse(ws, txt, resultLine, rowIndex, colIndex)
            if endLine in txt :
                readFlag = "False"
                if endFlag == "y" : write_txt += txt
                ws.cell(rowIndex, colIndex, write_txt)
                write_txt = newStartLine
            if readFlag == "True":
                write_txt += txt
            if not txt: break

        txt_f.close()

        colIndex += resultFlag

    # 스타일 설정
    SetStyle(resultFlag, ws, colIndex, rowIndex)

    # 파일 저장
    wb.save(resultFile)


def main():
    # 기존 파일 있으면 삭제
    RemoveFile()

    # 정보 입력
    pathDir, startLine, startFlag, newStartLine, endLine, endFlag, fileList, resultFlag, resultLine = InputInformation()

    # utf-8로 변환
    ConvertFile(pathDir, fileList)

    # 파싱 실행
    StartParse(pathDir, startLine, startFlag, newStartLine, endLine, endFlag, fileList, resultFlag, resultLine)

    # 종료
    print("\n*** 완료되었습니다. ***\n*** \"ParsingTool.exe\"이 있는 경로에 \"ParsingResult.xlsx\" 결과 파일이 생성됩니다. ***")
    time.sleep(5)


if __name__ == '__main__':
    main()
