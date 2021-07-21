# -*- coding: utf-8 -*-
# PYTHON 3.6, 3.7, 3.9

# pyinstaller --icon=pngegg.ico --onefile ParsingTool.py

# 개선할 점
# 1. default 문자열
# 2. 선택한 dir 출력
# 3. 현재상태
# 4. 양호취약 판별 로직 수정
# 5. 페이지 나누기 미리보기
# 6. utf-8 변환 에러 수정
# 7. 보고서 출력

import io, os, time, sys
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

global startString
startString =""

window=Tk()

#기본 설정
window.title("GUI Test")
window.geometry("600x400+700+250")  #너비X높이+X좌표+Y좌표
window.resizable(False, False)        #사이즈 변경 가능

def FolderOpen():
    global pathDir
    pathDir = filedialog.askdirectory(initialdir="/", title="Select Folder")

def close():
    window.quit()
    window.destroy()

#메뉴
menubar=Menu(window)

menu1=Menu(menubar, tearoff=0, selectcolor="red")
menu1.add_command(label="Open", command=FolderOpen)
menu1.add_separator()
menu1.add_command(label="Exit", command=close)
menubar.add_cascade(label="Menu", menu=menu1)

menu2=Menu(menubar, tearoff=0, selectcolor="red")
menu2.add_command(label="About...")
menubar.add_cascade(label="Help", menu=menu2)

window.config(menu=menubar)

def type(event):
    # 콤보박스2
    if Combx1.get() == "서버":
        values = ["Unix", "Window"]
    elif Combx1.get() == "WEB/WAS":
        values = ["Apache"]
    elif Combx1.get() == "DBMS":
        values = ["Oracle", "MySQL"]
    elif Combx1.get() == "PC":
        values = ["Windows", "MacOS"]
    else:
        values = ["해당없음"]
    Combx2 = ttk.Combobox(window, values=values)
    Combx2.current(0)
    Combx2.place(x=350, y=10)  # 콤보박스 배치

#콤보박스1 라벨
lbl1 = Label(window, text="대분류")
lbl1.place(x=10, y=10)

#콤보박스1
values1=["서버", "WEB/WAS", "DBMS", "PC", "네트워크 장비", "정보보호시스템 장비"]
Combx1=ttk.Combobox(window, values=values1)
Combx1.current(0)
Combx1.place(x=60, y=10) #콤보박스 배치
Combx1.bind('<<ComboboxSelected>>', type)

# 콤보박스2 라벨
lbl2 = Label(window, text="소분류")
lbl2.place(x=300, y=10)

# 시작 패턴 라벨
startLabel = Label(window, text="시작 문자열")
startLabel.place(x=10, y=50)

# 시작 패턴 입력
global startPattern
startPattern = Entry()
startPattern.place(x=100, y=50, width=410, heigh=30) #콤보박스 배치

# 마지막 패턴 라벨
endLabel = Label(window, text="마지막 문자열")
endLabel.place(x=10, y=100)

# 마지막 패턴 입력
global endPattern
endPattern = Entry(width=20)
endPattern.place(x=100, y=100, width=410, heigh=30) #콤보박스 배치

## 시작 문자열 추가 라벨
lbl3 = Label(window, text="현황 문자열 추가")
lbl3.place(x=10, y=150)

# 시작 문자열 추가
def StartStringY(event) :
    # 시작 문자열 추가 라벨
    label4 = Label(window, text="현황 문자열")
    label4.place(x=170, y=150)

    # 시작 문자열 입력
    StartStringEntry = Entry(width=20)
    StartStringEntry.place(x=210, y=150, width=300, heigh=30)  # 콤보박스 배치
    StartString = StartStringEntry.get()

# 시작 문자열 라디오 버튼
RadioVariety1=IntVar()

radio1=Radiobutton(window, text="Y", variable=RadioVariety1, value=1)
radio1.pack()
radio1.place(x=120, y=150)
radio1.bind('<Button-1>', StartStringY)

radio2=Radiobutton(window, text="N", variable=RadioVariety1, value=0)
radio2.pack()
radio2.place(x=120, y=180)

# 취약 양호 여부 파싱
###################

def RemoveFile() :
    if os.path.isfile("상세결과.xlsx"):
        removeFlag = input("\n*** 기존 파일(\"상세결과.xlsx\")이 존재합니다.\n*** 계속 진행하려면 삭제해야 합니다. 삭제하시겠습니까?\n  예:y, 아니오:n\n =>  ")
        if removeFlag == "y" :
            try:
                os.remove("상세결과.xlsx")
            except PermissionError:
                print("\n*** \"상세결과.xlsx\"파일을 닫은 후 재실행해주세요.")
                time.sleep(2)
                sys.exit()
            print("\n*** 기존 파일(\"상세결과.xlsx\")을 삭제하였습니다.")
        if removeFlag == "n":
            print("기존 파일(\"상세결과.xlsx\") 삭제 후 실행해야 합니다.")
            time.sleep(2)
            sys.exit()

def ConvertFile(pathDir):
    print(pathDir)
    global fileList
    fileList = os.listdir(pathDir)

    print("\n*** 파싱을 위해 대상 파일을 utf-8로 변환 실행")
    for fileName in fileList:
        fileDir = pathDir + '\\' + fileName
        try:
            file = io.open(fileDir, mode='r').read()
        except UnicodeError:
            continue
        io.open(fileDir, mode='w', encoding="utf-8").write(file)
    print("\n*** utf-8 포맷으로 변환 완료")

def SetStyle(ws, rowIndex):
    ScolIndex = 11
    SrowIndex = rowIndex

    # 행 높이 설정
    for row in range(3, rowIndex + 1):
        ws.row_dimensions[row].height = 18

    # 열 너비 설정
    ws.column_dimensions['J'].width = 47

    # 기본 틀 스타일 설정
    # 제목 글자
    ws.cell(1, 1).font = Font(bold=True, size=11, color='000000')

    # 표 제목 글자
    for r in range(1, ScolIndex+1):
        ws.cell(3, r).font = Font(bold=True, size=9, color='002060')
        ws.cell(3, r).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(3, r).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 표 내용 글자
    for i in range(4, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            ws.cell(i, s).font = Font(size=9, color='000000')


    # 테두리 설정
    # 표 내용 테두리
    for i in range(4, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            ws.cell(i, s).border = Border(left =Side(style='thin', color='D9D9D9'),
                                            right = Side(style='thin', color='D9D9D9'),
                                            top=Side(style='thin', color='D9D9D9'),
                                            bottom=Side(style='thin', color='D9D9D9'))
    # 표 제목 테두리
    for s in range(1, ScolIndex+1):
        ws.cell(3, s).border = Border(left =Side(style='thin', color='D9D9D9'),
                                        right = Side(style='thin', color='D9D9D9'),
                                        top=Side(style='medium', color='002060'),
                                        bottom=Side(style='thin', color='002060'))

def StartParse() :
    # 엑셀 시트 및 기본 틀 생성
    wb = Workbook()
    ws = wb.active

    ws.sheet_view.showGridLines = False

    #ws.sheet_properties.pageSetUpPr = PageSetupProperties(autoPageBreaks=True, fitToPage=True)
    #ws.sheet_properties.pageSetUpPr.autoPageBreaks=True

    ws.title = "상세결과"

    ws.cell(1, 1, "■ 상세결과")
    ws.cell(3, 1, "점검영역")
    ws.cell(3, 2, "CODE")
    ws.cell(3, 3, "점검항목")
    ws.cell(3, 4, "위험도")
    ws.cell(3, 5, "자산명")
    ws.cell(3, 6, "IP")
    ws.cell(3, 7, "분류")
    ws.cell(3, 8, "진단결과")
    ws.cell(3, 9, "조치결과")
    ws.cell(3, 10, "현재설정")
    ws.cell(3, 11, "비고")

    readFlag = "False"
    write_txt = startString
    rowIndex = 3
    colIndex = 10
    resultFile = "상세결과.xlsx"

    # 파일 하나씩 열어서 파싱
    for fileName in fileList:
        fileDir = pathDir + '\\' + fileName
        txt_f = open(fileDir, encoding="utf-8", mode="r")

        while True:
            txt = txt_f.readline()
            if startPattern.get() in txt :
                readFlag = "True"
                rowIndex += 1
                continue
            if endPattern.get() in txt :
                readFlag = "False"
                ws.cell(rowIndex, 5, fileName)
                ws.cell(rowIndex, 6, fileName)
                ws.cell(rowIndex, 7, fileName)
                ws.cell(rowIndex, colIndex, write_txt)
                ws.cell(rowIndex, colIndex).alignment = Alignment(wrap_text=True)
                write_txt = startString
            if readFlag == "True":
                write_txt += txt
            if not txt: break

        txt_f.close()

    # 스타일 설정
    SetStyle(ws, rowIndex)

    # 파일 저장
    wb.save(resultFile)

def ParsingMain():
    # 기존 파일 있으면 삭제
    RemoveFile()

    # utf-8로 변환
    ConvertFile(pathDir)

    # 파싱 실행
    StartParse()

    print("\n*** 완료 ***\n*** \"파싱 프로그램\"이 있는 경로에 \"상세결과.xlsx\" 결과 파일 생성 ***")


#버튼
btn1=Button(window, text="파싱", bg="yellow", fg="green", command=ParsingMain)
btn2=Button(window, text="보고서 생성", bg="yellow", fg="green")

btn1.place(x=120, y=320, width=120, height=40)
btn2.place(x=270, y=320, width=120, height=40)


def main() :
    window.mainloop()

if __name__ == '__main__':
    main()
